r"""
Migrate the client's BOM/purchasing .xlsx files into the operations DB.

Reads (default):
    01_Client_Analysis/client_data/EasySTimDataSBushColouredBeads.xlsx
    01_Client_Analysis/client_data/EasySTimDataSBushStandardBeads.xlsx
    01_Client_Analysis/client_data/EasyTimDataSBushFindingsEtc.xlsx

Decisions baked in:
- Skip `AGCCLRDBeads` and `MASTERBeadSheet` (client confirmed out of date).
- Skip `AGCFindingsMadeInHouse` (extender-chain sub-assembly handled as
  fractional metres of 1m chain in BOMs directly).
- Treat spreadsheet column `SBR00MULTI` as alias for the new code `SBA00MULTI`.
- FBA bundles (SBA02FBA / SBA04FBA / SBA06FBA) are auto-created with a flat
  BOM = sum of constituent SBA element BOMs for that palette.
- Idempotent. Default does NOT touch current_stock on re-run; pass
  --with-initial-stock on first real load to emit INITIAL_STOCK movements
  from the SOHAND UNIT column.

Usage:
    python manage.py migrate_bom_xlsx --dry-run
    python manage.py migrate_bom_xlsx --with-initial-stock
"""

from __future__ import annotations

import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from inventory.models import (
    BomLine,
    Brand,
    Product,
    ProductVariant,
    RawMaterial,
    StockMovement,
    Supplier,
    Variant,
)


# ---------------------------------------------------------------------------
# Reference data — locked with the client
# ---------------------------------------------------------------------------


BRANDS = {
    "SBR": {"name": "Sugar Bush", "description": "Beach Range — bikini brand / channel partner."},
    "SBA": {"name": "African Goddess", "description": "Multi-Wear Body Adornment range."},
}

VARIANTS = {
    # Sugar Bush Beach Range — 6 colour schemes
    "SBR01": ("SBR", "Tangerine & Orange"),
    "SBR02": ("SBR", "Coral & Magenta"),
    "SBR03": ("SBR", "Lime Green"),
    "SBR04": ("SBR", "Teal & Turquoise"),
    "SBR05": ("SBR", "Periwinkle & Powder Blue"),
    "SBR06": ("SBR", "Beige, White & Clear"),
    # African Goddess Multi-Wear — 3 named palettes
    "SBA02": ("SBA", "Mardi Gras — Amethyst, Coral, Magenta"),
    "SBA04": ("SBA", "Mystic Lagoon — Amazonite, Teal"),
    "SBA06": ("SBA", "Pure Sanctity — Beige, Clear & White Agate"),
}

# Product type codes (the "00" master-code) -> human names + brand
PRODUCTS = {
    "SBR00EARR":  ("SBR", "Beach Range Earrings (Pair)"),
    "SBR00DNL":   ("SBR", "Beach Range Double Necklace"),
    "SBR00DBL":   ("SBR", "Beach Range Double Bracelet"),
    "SBR00DWC":   ("SBR", "Beach Range Double Waist Chain"),
    "SBR00DANK":  ("SBR", "Beach Range Double Anklet"),
    "SBR00BBS":   ("SBR", "Beach Range Barefoot Beach Sandals (Pair)"),

    "SBA00EARR":  ("SBA", "Body Adornment Earrings (Pair)"),
    "SBA00NL":    ("SBA", "Body Adornment Necklace"),
    "SBA00BK":    ("SBA", "Body Adornment Back Piece"),
    "SBA00SN":    ("SBA", "Body Adornment Stanagra (Nipple) Nooses (Pair)"),
    "SBA00SL":    ("SBA", "Body Adornment Stanagra (Nipple) Loops (Pair)"),
    "SBA00WC":    ("SBA", "Body Adornment Waist Chain"),
    "SBA00MULTI": ("SBA", "Body Adornment Multi-wear (Head/Neck/Back/Shoulder/Hip)"),
    "SBA00FBA":   ("SBA", "Full Body Adornment Set"),
}

# Bundles: which ProductVariants make up a bundle ProductVariant
# Keyed by bundle product code; value = list of element product codes (per palette)
BUNDLES = {
    "SBA00FBA": ["SBA00EARR", "SBA00NL", "SBA00BK",
                 "SBA00SN", "SBA00SL", "SBA00WC"],
}

# Variants in the SBA brand that get FBA bundles
FBA_VARIANTS = ["SBA02", "SBA04", "SBA06"]

# Cross-brand variant mapping. The SBR coloured beads sheet groups beads by
# colour palette (SBR02 - Magenta & Coral / Amethyst Mardi Gras, etc.) but
# the same row carries BOM quantities for BOTH SBR* product columns
# (Sugar Bush) and SBA* product columns (African Goddess). When we're
# inside an SBR section but the BOM cell sits in an SBA product column,
# apply the line to the matching SBA variant rather than dropping it.
CROSS_BRAND_VARIANT_MAP = {
    "SBR02": "SBA02",  # Magenta & Coral / Amethyst Mardi Gras
    "SBR04": "SBA04",  # Teal & Turquoise / Mystic Lagoon
    "SBR06": "SBA06",  # Beige, White & Clear / Pure Sanctity
    # SBR01, SBR03, SBR05 — Sugar Bush only, no AG equivalent
}

# Column-name aliases (spreadsheet column -> canonical product code)
COLUMN_ALIASES = {
    "SBR00MULTI": "SBA00MULTI",
}

# Sheets to ignore entirely
IGNORED_SHEETS = {"AGCCLRDBeads", "MASTERBeadSheet", "AGCFindingsMadeInHouse",
                  "BeadPlotting"}


# Regex to spot a variant code in a "COLOUR COMBINATION" cell
VARIANT_CODE_RE = re.compile(r"^\s*(SB[RA]\d{2})\b", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------


class Command(BaseCommand):
    help = "Import African Goddess BOM data from the client's .xlsx files."

    def add_arguments(self, parser):
        parser.add_argument(
            "--data-dir",
            default=str(self._default_data_dir()),
            help="Folder containing the .xlsx files. Defaults to the knowledge-repo client_data folder.",
        )
        parser.add_argument("--dry-run", action="store_true",
                            help="Validate but rollback at the end.")
        parser.add_argument("--with-initial-stock", action="store_true",
                            help="Set current_stock from SOHAND UNIT and emit INITIAL_STOCK movements. "
                                 "Use only on first import; not idempotent on stock.")

    @staticmethod
    def _default_data_dir() -> Path:
        # v2: BEAD TOTALS rows now have correct =SUM(...) formulas. Originals
        # are preserved in ../client_data/ for diff/audit.
        base = Path(settings.BASE_DIR).parent
        candidate = base / "Inventory & Sales Management" / "01_Client_Analysis" / "client_data_v2"
        return candidate

    # -----------------------------------------------------------------------
    def handle(self, *args, **options):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise CommandError(
                "openpyxl is required. Install with: .venv\\Scripts\\python.exe -m pip install openpyxl"
            )

        data_dir = Path(options["data_dir"])
        if not data_dir.is_dir():
            raise CommandError(f"data dir not found: {data_dir}")

        files = sorted(data_dir.glob("*.xlsx"))
        if not files:
            raise CommandError(f"No .xlsx files found in {data_dir}")

        self.stdout.write(self.style.NOTICE(
            f"\n{'=' * 70}\n"
            f"{'DRY RUN' if options['dry_run'] else 'IMPORT'} from {data_dir}\n"
            f"{'=' * 70}"
        ))
        for f in files:
            self.stdout.write(f"  - {f.name}")
        self.stdout.write("")

        # Tally counters
        self.created = defaultdict(int)
        self.updated = defaultdict(int)
        self.skipped_rows = 0
        self.bom_lines_written = 0
        self.warnings = []

        with transaction.atomic():
            self._setup_reference_data()
            for f in files:
                self._process_file(f, with_initial_stock=options["with_initial_stock"])
            self._build_fba_bundles()

            if options["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("\nDRY RUN — rolled back."))

        self._report()

    # -----------------------------------------------------------------------
    # Reference data setup
    # -----------------------------------------------------------------------

    def _setup_reference_data(self) -> None:
        # Brands
        for code, info in BRANDS.items():
            obj, created = Brand.objects.update_or_create(
                code=code,
                defaults={"name": info["name"], "description": info["description"], "is_active": True},
            )
            self._tally("Brand", created)

        # Variants
        for code, (brand_code, name) in VARIANTS.items():
            brand = Brand.objects.get(code=brand_code)
            obj, created = Variant.objects.update_or_create(
                code=code,
                defaults={"name": name, "brand": brand, "is_active": True},
            )
            self._tally("Variant", created)

        # Products
        for code, (brand_code, name) in PRODUCTS.items():
            brand = Brand.objects.get(code=brand_code)
            obj, created = Product.objects.update_or_create(
                code=code,
                defaults={"name": name, "brand": brand,
                          "pillar": "BODY_ADORNMENTS", "is_active": True},
            )
            self._tally("Product", created)

        # ProductVariants — every (Product × applicable Variant) combination
        # SBR products → SBR01..SBR06
        # SBA element products + FBA → SBA02, SBA04, SBA06
        sbr_variants = [v for v in VARIANTS if v.startswith("SBR")]
        sba_variants = [v for v in VARIANTS if v.startswith("SBA")]
        for product_code, (brand_code, _) in PRODUCTS.items():
            applicable = sbr_variants if brand_code == "SBR" else sba_variants
            for variant_code in applicable:
                # Sellable SKU = variant_code + product type suffix (last 4-5 chars after SB?00)
                # e.g. SBR01 + EARR -> SBR01EARR ; SBA02 + FBA -> SBA02FBA
                suffix = product_code[5:]  # strip "SBR00" / "SBA00"
                sku = f"{variant_code}{suffix}"
                product = Product.objects.get(code=product_code)
                variant = Variant.objects.get(code=variant_code)
                obj, created = ProductVariant.objects.update_or_create(
                    product=product, variant=variant,
                    defaults={"sku": sku, "is_active": True},
                )
                self._tally("ProductVariant", created)

        # Default supplier
        Supplier.objects.update_or_create(name="Temu",
                                          defaults={"is_active": True,
                                                    "notes": "Auto-created by xlsx importer."})

    # -----------------------------------------------------------------------
    # Sheet processing
    # -----------------------------------------------------------------------

    def _process_file(self, path: Path, *, with_initial_stock: bool) -> None:
        import openpyxl
        self.stdout.write(self.style.NOTICE(f"\n--- {path.name}"))
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in IGNORED_SHEETS:
                self.stdout.write(f"  [skip] {sheet_name} (ignored)")
                continue
            ws = wb[sheet_name]
            self.stdout.write(f"  [read] {sheet_name}  ({ws.max_row} rows)")
            self._process_sheet(sheet_name, ws, with_initial_stock=with_initial_stock)

    def _process_sheet(self, sheet_name, ws, *, with_initial_stock: bool) -> None:
        # Load all rows into memory (sheets are small)
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if len(rows) < 3:
            return

        # Find header row: the row with the most non-empty string cells in the first 5
        header_idx = self._find_header_row(rows)
        header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        col_map = {name: idx for idx, name in enumerate(header) if name}

        # Detect product columns (any in our PRODUCTS dict OR a known alias)
        product_columns: dict[int, str] = {}
        for idx, name in enumerate(header):
            if not name:
                continue
            canonical = COLUMN_ALIASES.get(name, name)
            if canonical in PRODUCTS:
                product_columns[idx] = canonical

        if not product_columns:
            self.stdout.write(f"      (no product columns detected — skipping)")
            return

        # Walk data rows. Track current variant from COLOUR COMBINATION column.
        cc_idx = col_map.get("COLOUR COMBINATION")
        current_variant_code: Optional[str] = None
        for row_no in range(header_idx + 1, len(rows)):
            row = rows[row_no]
            # Update variant tracker
            if cc_idx is not None and cc_idx < len(row):
                cc_val = row[cc_idx]
                if cc_val:
                    m = VARIANT_CODE_RE.match(str(cc_val))
                    if m:
                        current_variant_code = m.group(1).upper()

            # Skip rows with no usable item ID
            sku_idx = col_map.get("CURRENT ITEM ID CODE")
            if sku_idx is None:
                continue
            raw_sku = row[sku_idx] if sku_idx < len(row) else None
            if not raw_sku:
                continue
            base_sku = str(raw_sku).strip()
            if not base_sku:
                continue

            # Disambiguate by size + colour: each spreadsheet row = its own
            # material. Supplier listing IDs cover multiple sizes; some size
            # buckets cover multiple colours. Treat each combo as distinct.
            def _tok(name: str) -> str:
                idx = col_map.get(name)
                if idx is None or idx >= len(row):
                    return ""
                v = row[idx]
                if v in (None, ""):
                    return ""
                return re.sub(r"[^a-zA-Z0-9]+", "_", str(v).strip()).strip("_")

            sku = base_sku
            for tok in (_tok("ITEM SIZE"), _tok("COLOUR:") or _tok("COLOUR")):
                if tok:
                    sku = f"{sku}_{tok}"
            sku = sku[:64]

            # ----- Material upsert -----
            material = self._upsert_material(row, col_map, sku)
            if material is None:
                self.skipped_rows += 1
                continue

            # ----- Initial stock (one-shot) -----
            if with_initial_stock and material.current_stock == Decimal("0"):
                so_idx = col_map.get("SOHAND UNIT")
                so_val = self._dec(row[so_idx]) if so_idx is not None and so_idx < len(row) else None
                if so_val and so_val > 0:
                    material.current_stock = so_val
                    material.save(update_fields=["current_stock", "updated_at"])
                    StockMovement.objects.create(
                        raw_material=material, delta=so_val, reason="INITIAL_STOCK",
                        note=f"Imported from {sheet_name} (SOHAND UNIT)",
                    )

            # ----- BOM lines -----
            for col_idx, product_code in product_columns.items():
                if col_idx >= len(row):
                    continue
                qty = self._dec(row[col_idx])
                if not qty or qty <= 0:
                    continue

                product = Product.objects.get(code=product_code)
                applicable_variants = self._applicable_variants(product, current_variant_code)
                for variant_code in applicable_variants:
                    pv = ProductVariant.objects.filter(
                        product=product, variant__code=variant_code,
                    ).first()
                    if not pv:
                        self.warnings.append(
                            f"  ! No ProductVariant for {variant_code} x {product_code} "
                            f"(row {row_no + 1})"
                        )
                        continue
                    BomLine.objects.update_or_create(
                        product_variant=pv, raw_material=material,
                        defaults={"quantity": qty},
                    )
                    self.bom_lines_written += 1

    @staticmethod
    def _applicable_variants(product: Product, current_variant_code: Optional[str]) -> list[str]:
        """Decide which variant(s) this BOM line applies to.

        - Same-brand match: if the row's variant matches the product's brand
          (e.g. inside SBR02, applying to an SBR product) — use the row variant.
        - Cross-brand match: SBR sections in the coloured-beads sheet carry
          BOM quantities for SBA columns too (the colour palette is shared
          with the matching African Goddess variant). Map SBR -> SBA via
          CROSS_BRAND_VARIANT_MAP and apply if a mapping exists.
        - No variant tracked (findings, standard beads) — cross-cutting:
          apply to ALL variants of this product's brand.
        """
        if current_variant_code:
            v_brand = current_variant_code[:3]  # SBR or SBA
            if v_brand == product.brand.code:
                return [current_variant_code]
            # Cross-brand: SBR row -> matching SBA variant if mapped
            if v_brand == "SBR" and product.brand.code == "SBA":
                mapped = CROSS_BRAND_VARIANT_MAP.get(current_variant_code)
                if mapped:
                    return [mapped]
            return []
        # No variant tracker → cross-cutting. Apply to all variants of this
        # product's brand.
        return [v for v, (b, _) in VARIANTS.items() if b == product.brand.code]

    @staticmethod
    def _find_header_row(rows: list[list]) -> int:
        for i, row in enumerate(rows[:5]):
            non_empty = [c for c in row if c not in (None, "")]
            string_like = [c for c in non_empty if isinstance(c, str)]
            if non_empty and len(string_like) >= max(2, len(non_empty) - 1):
                return i
        return 0

    # -----------------------------------------------------------------------
    # Material upsert
    # -----------------------------------------------------------------------

    def _upsert_material(self, row, col_map, sku: str) -> Optional[RawMaterial]:
        def cell(name: str) -> Optional[str]:
            idx = col_map.get(name)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v not in (None, "") else None

        # Build composite name
        descr = cell("DESCRIPTION") or ""
        colour = cell("COLOUR:") or cell("COLOUR") or ""
        size = cell("ITEM SIZE") or ""
        shape = cell("SHAPE") or ""
        finish = cell("FINISH") or ""
        sub_brand = cell("BRAND") or ""
        if sub_brand.lower() in ("none", "n/a"):
            sub_brand = ""
        alt_id = cell("ALTERNATIVE ITEM ID CODE") or ""
        notes = cell("NOTES") or ""

        # Name: use DESCRIPTION if present, else build from parts
        name = descr or " ".join(p for p in (size, colour, shape) if p) or sku
        # Truncate name field to model max
        name = name[:255]

        unit = self._guess_unit(size, descr, shape)

        pack_size = self._dec(cell("PACK SIZE")) or Decimal("1")
        pack_cost = self._dec(cell("TEMU PACK COST")) or Decimal("0")
        unit_cost = self._dec(cell("FINAL PER UNIT PRICE")) or Decimal("0")
        if unit_cost == 0 and pack_size > 0 and pack_cost > 0:
            unit_cost = (pack_cost / pack_size).quantize(Decimal("0.0001"))
        reorder_point = self._dec(cell("UNIT QTY FOR BASE STOCK LEVEL")) or Decimal("0")

        supplier_name = (cell("SUPPLIER") or "").strip()
        supplier = None
        if supplier_name:
            supplier, _ = Supplier.objects.get_or_create(name=supplier_name)

        defaults = {
            "name": name[:255],
            "internal_id_code": (cell("INTERNAL ID CODE") or "")[:32],
            "alternative_id_code": alt_id[:64],
            "item_size": size[:64],
            "colour": colour[:128],
            "finish": finish[:128],
            "shape": shape[:128],
            "description": descr,
            "sub_brand": sub_brand[:64],
            "unit": unit,
            "reorder_point": reorder_point,
            "reorder_quantity": pack_size,  # default reorder qty = one pack
            "pack_size": pack_size,
            "last_paid_pack_cost": pack_cost,
            "last_paid_unit_cost": unit_cost,
            "preferred_supplier": supplier,
            "notes": notes[:5000] if notes else "",
            "is_active": True,
        }

        try:
            obj, created = RawMaterial.objects.update_or_create(sku=sku, defaults=defaults)
            self._tally("RawMaterial", created)
            return obj
        except Exception as exc:
            self.warnings.append(f"  ! could not upsert material sku={sku}: {exc}")
            return None

    @staticmethod
    def _guess_unit(size, descr, shape) -> str:
        s = " ".join(filter(None, (size, descr, shape))).lower()
        if "metre" in s or " m " in s or "1m" in s or "lengths" in s or "chain" in s and "extender" not in s:
            return "metre"
        if "gram" in s or " g " in s:
            return "gram"
        return "piece"

    # -----------------------------------------------------------------------
    # FBA bundle BOMs (sum of constituents)
    # -----------------------------------------------------------------------

    def _build_fba_bundles(self) -> None:
        from collections import defaultdict
        for variant_code in FBA_VARIANTS:
            try:
                fba_pv = ProductVariant.objects.get(
                    product__code="SBA00FBA", variant__code=variant_code,
                )
            except ProductVariant.DoesNotExist:
                continue
            # Wipe existing FBA BOM lines then rebuild as sum of element BOMs
            BomLine.objects.filter(product_variant=fba_pv).delete()
            totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
            for element_code in BUNDLES["SBA00FBA"]:
                try:
                    element_pv = ProductVariant.objects.get(
                        product__code=element_code, variant__code=variant_code,
                    )
                except ProductVariant.DoesNotExist:
                    continue
                for line in element_pv.bom_lines.all():
                    totals[line.raw_material_id] += line.quantity
            for material_id, qty in totals.items():
                BomLine.objects.create(
                    product_variant=fba_pv,
                    raw_material_id=material_id,
                    quantity=qty,
                    notes="Auto-summed from FBA constituents",
                )
                self.bom_lines_written += 1
            self.stdout.write(f"  [bundle] {fba_pv.sku}: BOM = sum of "
                              f"{len(BUNDLES['SBA00FBA'])} elements ({len(totals)} materials)")

    # -----------------------------------------------------------------------
    # Helpers
    # -----------------------------------------------------------------------

    @staticmethod
    def _dec(value) -> Optional[Decimal]:
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def _tally(self, model_name: str, created: bool) -> None:
        if created:
            self.created[model_name] += 1
        else:
            self.updated[model_name] += 1

    # -----------------------------------------------------------------------
    def _report(self) -> None:
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== Summary ==="))
        all_models = sorted(set(self.created) | set(self.updated))
        for m in all_models:
            self.stdout.write(f"  {m}: {self.created[m]} created, {self.updated[m]} updated")
        self.stdout.write(f"  BomLines written: {self.bom_lines_written}")
        self.stdout.write(f"  Skipped rows: {self.skipped_rows}")
        if self.warnings:
            self.stdout.write(self.style.WARNING(f"\n{len(self.warnings)} warning(s):"))
            for w in self.warnings[:30]:
                self.stdout.write(w)
            if len(self.warnings) > 30:
                self.stdout.write(f"  ... and {len(self.warnings) - 30} more")
