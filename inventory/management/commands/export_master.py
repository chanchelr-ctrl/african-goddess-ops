"""Export the canonical master data .xlsx — three sheets:

  MaterialMaster — one row per RawMaterial (operational state, stock last)
  ProductSpec    — one row per BomLine (design-controlled spec)
  ChangeLog      — full DataChangeLog (audit trail)

The DB is the system of record. This export is a point-in-time snapshot
generated on demand. import_master can read this same file back in.

Usage:
  python manage.py export_master
  python manage.py export_master --out path/to/file.xlsx
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.core.management.base import BaseCommand

from inventory.models import BomLine, DataChangeLog, RawMaterial


CURRENCY_FMT = '"R "#,##0.00'
WHOLE_NUMBER_FMT = "#,##0"


# ---------------------------------------------------------------------------
# Header schemas (kept stable so import_master can mirror them)
# ---------------------------------------------------------------------------


MATERIAL_MASTER_HEADERS = [
    "SKU",
    "Supplier ID (Current)",
    "Alternative ID",
    "Internal ID",
    "Name",
    "Description",
    "Item Size",
    "Colour",
    "Finish",
    "Shape",
    "Sub-brand",
    "Unit",
    "Pack Size",
    "Pack Cost (ZAR)",
    "Import Duties / Pack",
    "Unit Cost (ZAR)",
    "Supplier",
    "Reorder Point",
    "Reorder Quantity",
    "Active?",
    "Notes",
    "Stock on Hand",
    "ZAR Value of Stock on Hand",
]

PRODUCT_SPEC_HEADERS = [
    "PV SKU",
    "Product Code",
    "Product Name",
    "Variant Code",
    "Variant Name",
    "Brand",
    "Retail Price (ZAR)",
    "Material SKU",
    "Material Name",
    "Item Size",
    "Colour",
    "Shape",
    "BOM Quantity",
    "Notes",
]

CHANGELOG_HEADERS = [
    "Timestamp",
    "User",
    "Action",
    "Model",
    "SKU/Code",
    "Object PK",
    "Field",
    "Old Value",
    "New Value",
    "Note",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _supplier_id_from_sku(sku: str) -> str:
    return sku.split("_", 1)[0]


def _style_header_row(ws):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), max_width)


def _fmt_columns(ws, headers, cols, fmt):
    for label in cols:
        if label not in headers:
            continue
        c = headers.index(label) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = fmt


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_material_master(wb):
    ws = wb.active
    ws.title = "MaterialMaster"
    ws.append(MATERIAL_MASTER_HEADERS)

    unit_cost_col = MATERIAL_MASTER_HEADERS.index("Unit Cost (ZAR)") + 1
    stock_col = MATERIAL_MASTER_HEADERS.index("Stock on Hand") + 1
    value_col = MATERIAL_MASTER_HEADERS.index("ZAR Value of Stock on Hand") + 1
    unit_cost_letter = get_column_letter(unit_cost_col)
    stock_letter = get_column_letter(stock_col)

    for m in RawMaterial.objects.select_related("preferred_supplier").order_by("sku"):
        ws.append([
            m.sku,
            _supplier_id_from_sku(m.sku),
            m.alternative_id_code or "",
            m.internal_id_code or "",
            m.name,
            m.description or "",
            m.item_size or "",
            m.colour or "",
            m.finish or "",
            m.shape or "",
            m.sub_brand or "",
            m.get_unit_display(),
            float(m.pack_size),
            float(m.last_paid_pack_cost),
            float(m.import_duties_per_pack),
            float(m.last_paid_unit_cost),
            m.preferred_supplier.name if m.preferred_supplier else "",
            float(m.reorder_point),
            float(m.reorder_quantity),
            "Yes" if m.is_active else "No",
            m.notes or "",
            float(m.current_stock),
            None,  # filled with formula below
        ])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=value_col).value = f"={stock_letter}{r}*{unit_cost_letter}{r}"

    _style_header_row(ws)
    _fmt_columns(ws, MATERIAL_MASTER_HEADERS,
                 ["Pack Size", "Reorder Point", "Reorder Quantity", "Stock on Hand"],
                 WHOLE_NUMBER_FMT)
    _fmt_columns(ws, MATERIAL_MASTER_HEADERS,
                 ["Pack Cost (ZAR)", "Import Duties / Pack", "Unit Cost (ZAR)",
                  "ZAR Value of Stock on Hand"],
                 CURRENCY_FMT)
    ws.freeze_panes = "B2"
    _autosize(ws)


def _build_product_spec(wb):
    ws = wb.create_sheet("ProductSpec")
    ws.append(PRODUCT_SPEC_HEADERS)

    qs = (BomLine.objects
          .select_related("raw_material",
                          "product_variant",
                          "product_variant__product",
                          "product_variant__product__brand",
                          "product_variant__variant")
          .order_by("product_variant__product__brand__code",
                    "product_variant__variant__code",
                    "product_variant__product__code",
                    "raw_material__sku"))

    for line in qs:
        m = line.raw_material
        pv = line.product_variant
        ws.append([
            pv.sku,
            pv.product.code,
            pv.product.name,
            pv.variant.code,
            pv.variant.name,
            pv.product.brand.name,
            float(pv.effective_retail_price_zar),
            m.sku,
            m.name,
            m.item_size or "",
            m.colour or "",
            m.shape or "",
            float(line.quantity),
            line.notes or "",
        ])

    _style_header_row(ws)
    _fmt_columns(ws, PRODUCT_SPEC_HEADERS, ["BOM Quantity"], WHOLE_NUMBER_FMT)
    _fmt_columns(ws, PRODUCT_SPEC_HEADERS, ["Retail Price (ZAR)"], CURRENCY_FMT)
    ws.freeze_panes = "B2"
    _autosize(ws)


def _build_changelog(wb):
    ws = wb.create_sheet("ChangeLog")
    ws.append(CHANGELOG_HEADERS)

    qs = DataChangeLog.objects.select_related("user").order_by("-timestamp")
    for e in qs:
        ws.append([
            e.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            e.user.username if e.user else "system",
            e.action,
            e.model_name,
            e.sku or "",
            e.object_pk,
            e.field or "",
            e.old_value,
            e.new_value,
            e.note,
        ])

    _style_header_row(ws)
    ws.freeze_panes = "B2"
    _autosize(ws)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_master_workbook() -> openpyxl.Workbook:
    """Build and return the master workbook in memory. Used by the web view
    too, not just the management command."""
    wb = openpyxl.Workbook()
    _build_material_master(wb)
    _build_product_spec(wb)
    _build_changelog(wb)
    wb.active = 0
    return wb


def export_to_bytes() -> bytes:
    wb = build_master_workbook()
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------


class Command(BaseCommand):
    help = "Export master data + spec + change log to an .xlsx file."

    def add_arguments(self, parser):
        parser.add_argument(
            "--out",
            default=None,
            help="Output path. Defaults to ../client_data_v5/MasterData_v5.xlsx",
        )

    def handle(self, *args, **opts):
        out = opts["out"]
        if not out:
            base = Path(settings.BASE_DIR).parent
            out = base / "Inventory & Sales Management" / "01_Client_Analysis" / \
                  "client_data_v5" / "MasterData_v5.xlsx"
        out = Path(out)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = build_master_workbook()
        wb.save(out)

        for sn in ["MaterialMaster", "ProductSpec", "ChangeLog"]:
            ws = wb[sn]
            self.stdout.write(f"  {sn}: {ws.max_row - 1} rows x {ws.max_column} cols")
        self.stdout.write(self.style.SUCCESS(f"\nSaved -> {out}"))
