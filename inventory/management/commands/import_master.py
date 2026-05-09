"""Import the canonical master data .xlsx — reads MaterialMaster + ProductSpec
sheets and updates the DB.

Round-trips the schema written by export_master. Additive by default:
materials and BOM lines present in the file are upserted; rows missing from
the file are left alone (no auto-delete). Use --prune for stricter mode.

Behaviour:
  - MaterialMaster: matched by SKU. Existing materials updated; new SKUs
    created. Stock-on-Hand is treated as the new current_stock; we emit
    an ADJUSTMENT StockMovement to keep the audit log reconciled.
  - ProductSpec:    matched by (PV SKU, Material SKU). PVs and Materials
    must already exist (they are NOT auto-created from this sheet).
  - ChangeLog sheet: ignored on import (it's an output-only audit trail).

Usage:
  python manage.py import_master <path>
  python manage.py import_master <path> --dry-run
  python manage.py import_master <path> --prune
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from inventory.models import (
    BomLine,
    ProductVariant,
    RawMaterial,
    StockMovement,
    Supplier,
)


UNIT_REVERSE_MAP = {
    "pieces": "piece",
    "grams": "gram",
    "kilograms": "kg",
    "metres": "metre",
    "centimetres": "cm",
    "millilitres": "ml",
    "litres": "litre",
    "strands": "strand",
    "packs": "pack",
    "sets": "set",
    "other": "other",
}


def _to_decimal(v) -> Decimal | None:
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def _to_str(v) -> str:
    if v in (None, ""):
        return ""
    return str(v).strip()


class Command(BaseCommand):
    help = "Import master data + product spec from a .xlsx file."

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to MasterData_v?.xlsx")
        parser.add_argument("--dry-run", action="store_true",
                            help="Validate only; rollback at the end.")
        parser.add_argument(
            "--prune", action="store_true",
            help="Delete BomLines and Materials not present in the file. "
                 "Use with care.",
        )

    def handle(self, *args, **opts):
        path = Path(opts["path"])
        if not path.is_file():
            raise CommandError(f"File not found: {path}")

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as e:
            raise CommandError(f"Could not open workbook: {e}")

        if "MaterialMaster" not in wb.sheetnames:
            raise CommandError("Workbook is missing required sheet: MaterialMaster")
        if "ProductSpec" not in wb.sheetnames:
            raise CommandError("Workbook is missing required sheet: ProductSpec")

        self._counts = {
            "materials_created": 0, "materials_updated": 0,
            "stock_adjustments": 0,
            "boms_created": 0, "boms_updated": 0, "boms_unchanged": 0,
            "boms_skipped": 0,
            "materials_pruned": 0, "boms_pruned": 0,
        }
        self._warnings: list[str] = []

        with transaction.atomic():
            seen_material_skus = self._import_materials(wb["MaterialMaster"])
            seen_bom_keys = self._import_specs(wb["ProductSpec"])

            if opts["prune"]:
                self._prune_boms(seen_bom_keys)
                self._prune_materials(seen_material_skus)

            if opts["dry_run"]:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("\nDRY RUN — rolled back."))

        self._report()

    # -----------------------------------------------------------------------

    def _read_sheet(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        return header, rows[1:]

    def _idx(self, header, name):
        return header.index(name) if name in header else None

    # -----------------------------------------------------------------------
    # MaterialMaster
    # -----------------------------------------------------------------------

    def _import_materials(self, ws) -> set[str]:
        header, data = self._read_sheet(ws)

        I = lambda n: self._idx(header, n)  # noqa: E731
        col_sku        = I("SKU")
        col_alt_id     = I("Alternative ID")
        col_internal   = I("Internal ID")
        col_name       = I("Name")
        col_descr      = I("Description")
        col_size       = I("Item Size")
        col_colour     = I("Colour")
        col_finish     = I("Finish")
        col_shape      = I("Shape")
        col_subbrand   = I("Sub-brand")
        col_unit       = I("Unit")
        col_pack_size  = I("Pack Size")
        col_pack_cost  = I("Pack Cost (ZAR)")
        col_duties     = I("Import Duties / Pack")
        col_unit_cost  = I("Unit Cost (ZAR)")
        col_supplier   = I("Supplier")
        col_reorder    = I("Reorder Point")
        col_reorder_q  = I("Reorder Quantity")
        col_active     = I("Active?")
        col_notes      = I("Notes")
        col_stock      = I("Stock on Hand")

        if col_sku is None:
            raise CommandError("MaterialMaster sheet missing SKU column")

        seen = set()

        for row in data:
            sku = _to_str(row[col_sku]) if col_sku < len(row) else ""
            if not sku:
                continue
            seen.add(sku)

            supplier = None
            sup_name = _to_str(row[col_supplier]) if col_supplier is not None else ""
            if sup_name:
                supplier, _ = Supplier.objects.get_or_create(name=sup_name)

            unit_display = _to_str(row[col_unit]) if col_unit is not None else ""
            unit = UNIT_REVERSE_MAP.get(unit_display.lower(), "piece")

            new_stock = _to_decimal(row[col_stock]) if col_stock is not None else None
            new_stock = new_stock if new_stock is not None else Decimal("0")

            defaults = {
                "alternative_id_code": _to_str(row[col_alt_id])[:64] if col_alt_id is not None else "",
                "internal_id_code":    _to_str(row[col_internal])[:32] if col_internal is not None else "",
                "name":        _to_str(row[col_name])[:255] if col_name is not None else sku,
                "description": _to_str(row[col_descr]) if col_descr is not None else "",
                "item_size":   _to_str(row[col_size])[:64] if col_size is not None else "",
                "colour":      _to_str(row[col_colour])[:128] if col_colour is not None else "",
                "finish":      _to_str(row[col_finish])[:128] if col_finish is not None else "",
                "shape":       _to_str(row[col_shape])[:128] if col_shape is not None else "",
                "sub_brand":   _to_str(row[col_subbrand])[:64] if col_subbrand is not None else "",
                "unit":        unit,
                "pack_size":            _to_decimal(row[col_pack_size]) or Decimal("1") if col_pack_size is not None else Decimal("1"),
                "last_paid_pack_cost":  _to_decimal(row[col_pack_cost]) or Decimal("0") if col_pack_cost is not None else Decimal("0"),
                "import_duties_per_pack": _to_decimal(row[col_duties]) or Decimal("0") if col_duties is not None else Decimal("0"),
                "last_paid_unit_cost":  _to_decimal(row[col_unit_cost]) or Decimal("0") if col_unit_cost is not None else Decimal("0"),
                "preferred_supplier":   supplier,
                "reorder_point":        _to_decimal(row[col_reorder]) or Decimal("0") if col_reorder is not None else Decimal("0"),
                "reorder_quantity":     _to_decimal(row[col_reorder_q]) or Decimal("0") if col_reorder_q is not None else Decimal("0"),
                "is_active": (_to_str(row[col_active]).lower() != "no") if col_active is not None else True,
                "notes":     _to_str(row[col_notes]) if col_notes is not None else "",
            }

            obj, created = RawMaterial.objects.update_or_create(sku=sku, defaults=defaults)
            if created:
                self._counts["materials_created"] += 1
                # New material: set its stock and emit INITIAL_STOCK if non-zero
                if new_stock != 0:
                    obj.current_stock = new_stock
                    obj.save(update_fields=["current_stock", "updated_at"])
                    StockMovement.objects.create(
                        raw_material=obj, delta=new_stock, reason="INITIAL_STOCK",
                        note="Imported from MasterData",
                    )
            else:
                self._counts["materials_updated"] += 1
                # Existing material: if stock differs, emit an ADJUSTMENT for the delta
                delta = new_stock - obj.current_stock
                if delta != 0:
                    obj.current_stock = new_stock
                    obj.save(update_fields=["current_stock", "updated_at"])
                    StockMovement.objects.create(
                        raw_material=obj, delta=delta, reason="ADJUSTMENT",
                        note="Stock adjusted by master-data import",
                    )
                    self._counts["stock_adjustments"] += 1

        return seen

    # -----------------------------------------------------------------------
    # ProductSpec
    # -----------------------------------------------------------------------

    def _import_specs(self, ws) -> set[tuple[int, int]]:
        header, data = self._read_sheet(ws)

        I = lambda n: self._idx(header, n)  # noqa: E731
        col_pv     = I("PV SKU")
        col_msku   = I("Material SKU")
        col_qty    = I("BOM Quantity")
        col_notes  = I("Notes")
        if col_pv is None or col_msku is None or col_qty is None:
            raise CommandError(
                "ProductSpec sheet missing one of: PV SKU, Material SKU, BOM Quantity"
            )

        seen: set[tuple[int, int]] = set()
        pv_cache: dict[str, ProductVariant | None] = {}
        m_cache: dict[str, RawMaterial | None] = {}

        for row in data:
            pv_sku = _to_str(row[col_pv]) if col_pv < len(row) else ""
            m_sku  = _to_str(row[col_msku]) if col_msku < len(row) else ""
            qty    = _to_decimal(row[col_qty]) if col_qty < len(row) else None
            notes  = _to_str(row[col_notes]) if col_notes is not None and col_notes < len(row) else ""

            if not pv_sku or not m_sku or qty is None:
                continue
            if qty <= 0:
                # Treat zero/negative as "not in BOM" — skip
                self._counts["boms_skipped"] += 1
                continue

            if pv_sku not in pv_cache:
                pv_cache[pv_sku] = ProductVariant.objects.filter(sku=pv_sku).first()
            pv = pv_cache[pv_sku]
            if not pv:
                self._warnings.append(f"  ! Unknown PV SKU: {pv_sku} (row skipped)")
                self._counts["boms_skipped"] += 1
                continue

            if m_sku not in m_cache:
                m_cache[m_sku] = RawMaterial.objects.filter(sku=m_sku).first()
            m = m_cache[m_sku]
            if not m:
                self._warnings.append(f"  ! Unknown Material SKU: {m_sku} (row skipped)")
                self._counts["boms_skipped"] += 1
                continue

            existing = BomLine.objects.filter(product_variant=pv, raw_material=m).first()
            if existing:
                if existing.quantity == qty and existing.notes == notes:
                    self._counts["boms_unchanged"] += 1
                else:
                    existing.quantity = qty
                    existing.notes = notes
                    existing.save()
                    self._counts["boms_updated"] += 1
            else:
                BomLine.objects.create(
                    product_variant=pv, raw_material=m,
                    quantity=qty, notes=notes,
                )
                self._counts["boms_created"] += 1

            seen.add((pv.id, m.id))

        return seen

    # -----------------------------------------------------------------------
    # Prune
    # -----------------------------------------------------------------------

    def _prune_boms(self, seen: set[tuple[int, int]]):
        for line in BomLine.objects.all():
            if (line.product_variant_id, line.raw_material_id) not in seen:
                line.delete()
                self._counts["boms_pruned"] += 1

    def _prune_materials(self, seen: set[str]):
        for m in RawMaterial.objects.all():
            if m.sku not in seen:
                # Soft-prune: deactivate rather than delete (FK risk via BomLine,
                # POs, etc.). Pruned materials become is_active=False.
                if m.is_active:
                    m.is_active = False
                    m.save(update_fields=["is_active", "updated_at"])
                    self._counts["materials_pruned"] += 1

    # -----------------------------------------------------------------------

    def _report(self):
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=== import_master summary ==="))
        for k, v in self._counts.items():
            self.stdout.write(f"  {k}: {v}")
        if self._warnings:
            self.stdout.write(self.style.WARNING(f"\n{len(self._warnings)} warning(s):"))
            for w in self._warnings[:30]:
                self.stdout.write(w)
            if len(self._warnings) > 30:
                self.stdout.write(f"  ... and {len(self._warnings) - 30} more")
