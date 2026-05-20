"""
Compare two MasterData workbooks (xlsx-only, no DB).

Use when the client returns an edited workbook and we want to see exactly
what changed before importing.

For v6-1: the client copied ProductSpec into a new sheet "TUSHY" and updated
TUSHY, leaving the original ProductSpec stale. So for v6 the BOM source is
TUSHY, not ProductSpec.

Usage (from project root):
    .\\.venv\\Scripts\\python.exe scripts\\_compare_master_v5_v6.py
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

CLIENT_DIR = Path(r"C:\Users\chanr\Desktop\Inventory & Sales Management\01_Client_Analysis")
OLD_XLSX = CLIENT_DIR / "client_data_v5" / "MasterData_v5.xlsx"
NEW_XLSX = CLIENT_DIR / "client_data_v6" / "MasterData_v6-1.xlsx"

OLD_BOM_SHEET = "ProductSpec"
NEW_BOM_SHEET = "TUSHY"  # client's worked-on copy supersedes ProductSpec in v6


def _norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _num(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def read_sheet(path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return None, []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [_norm(c) for c in rows[0]]
    data = []
    for r in rows[1:]:
        if all(c is None or _norm(c) == "" for c in r):
            continue
        data.append(r)
    return header, data


def row_to_dict(header, row):
    return {header[i]: row[i] if i < len(row) else None for i in range(len(header))}


MATERIAL_FIELDS = [
    "Alternative ID", "Name", "Description", "Item Size", "Colour",
    "Finish", "Shape", "Sub-brand", "Unit", "Pack Size",
    "Pack Cost (ZAR)", "Import Duties / Pack", "Unit Cost (ZAR)",
    "Supplier", "Reorder Point", "Reorder Quantity", "Active?",
    "Notes", "Stock on Hand",
]
NUMERIC_FIELDS = {
    "Pack Size", "Pack Cost (ZAR)", "Import Duties / Pack", "Unit Cost (ZAR)",
    "Reorder Point", "Reorder Quantity", "Stock on Hand",
}


def compare_materials():
    print("=" * 70)
    print("MATERIAL MASTER")
    print("=" * 70)

    old_h, old_d = read_sheet(OLD_XLSX, "MaterialMaster")
    new_h, new_d = read_sheet(NEW_XLSX, "MaterialMaster")

    if old_h != new_h:
        added = [c for c in new_h if c not in old_h]
        removed = [c for c in old_h if c not in new_h]
        if added or removed:
            print(f"\n!! Header changed.  added: {added}  removed: {removed}")
        else:
            print(f"\n!! Column order changed (same set).")

    old_by_sku = {_norm(row_to_dict(old_h, r).get("SKU")): row_to_dict(old_h, r) for r in old_d}
    new_by_sku = {_norm(row_to_dict(new_h, r).get("SKU")): row_to_dict(new_h, r) for r in new_d}

    old_skus = set(old_by_sku) - {""}
    new_skus = set(new_by_sku) - {""}

    added_skus = sorted(new_skus - old_skus)
    removed_skus = sorted(old_skus - new_skus)
    common = sorted(old_skus & new_skus)

    print(f"\nRow counts:  v5 = {len(old_skus)}   v6 = {len(new_skus)}   common = {len(common)}")
    print(f"  NEW SKUs  (in v6, not in v5): {len(added_skus)}")
    print(f"  REMOVED   (in v5, not in v6): {len(removed_skus)}")

    if added_skus:
        print("\n--- NEW materials ---")
        for sku in added_skus:
            r = new_by_sku[sku]
            print(f"  + {sku}  |  {_norm(r.get('Name'))}  |  stock={_norm(r.get('Stock on Hand'))}  |  R{_norm(r.get('Unit Cost (ZAR)'))}/unit")

    if removed_skus:
        print("\n--- REMOVED materials ---")
        for sku in removed_skus:
            r = old_by_sku[sku]
            print(f"  - {sku}  |  {_norm(r.get('Name'))}  |  was-stock={_norm(r.get('Stock on Hand'))}")

    changed = []
    for sku in common:
        o = old_by_sku[sku]
        n = new_by_sku[sku]
        diffs = []
        for f in MATERIAL_FIELDS:
            ov, nv = o.get(f), n.get(f)
            if f in NUMERIC_FIELDS:
                od, nd = _num(ov), _num(nv)
                if od != nd:
                    diffs.append((f, _norm(ov), _norm(nv)))
            else:
                if _norm(ov) != _norm(nv):
                    diffs.append((f, _norm(ov), _norm(nv)))
        if diffs:
            changed.append((sku, _norm(n.get("Name")), diffs))

    print(f"\n  CHANGED materials: {len(changed)}")

    if changed:
        from collections import Counter
        bucket = Counter()
        for _, _, diffs in changed:
            for f, _o, _n in diffs:
                bucket[f] += 1
        print("\n  Changes by field:")
        for f, c in bucket.most_common():
            print(f"    {f:28s} {c}")


def compare_bom():
    print("\n" + "=" * 70)
    print(f"BOM  ({OLD_BOM_SHEET} in v5  vs  {NEW_BOM_SHEET} in v6)")
    print("=" * 70)

    old_h, old_d = read_sheet(OLD_XLSX, OLD_BOM_SHEET)
    new_h, new_d = read_sheet(NEW_XLSX, NEW_BOM_SHEET)

    def build_map(h, d):
        m = {}
        for r in d:
            row = row_to_dict(h, r)
            pv = _norm(row.get("PV SKU"))
            ms = _norm(row.get("Material SKU"))
            if not pv or not ms:
                continue
            m[(pv, ms)] = {
                "qty":   _num(row.get("BOM Quantity")),
                "notes": _norm(row.get("Notes")),
            }
        return m

    old_m = build_map(old_h, old_d)
    new_m = build_map(new_h, new_d)

    added = sorted(set(new_m) - set(old_m))
    removed = sorted(set(old_m) - set(new_m))
    common = sorted(set(old_m) & set(new_m))

    print(f"\nBOM line counts: v5 = {len(old_m)}   v6 = {len(new_m)}   common = {len(common)}")
    print(f"  NEW lines     (in v6, not in v5): {len(added)}")
    print(f"  REMOVED lines (in v5, not in v6): {len(removed)}")

    qty_changes = []
    note_changes = []
    for k in common:
        o = old_m[k]
        n = new_m[k]
        if o["qty"] != n["qty"]:
            qty_changes.append((k, o["qty"], n["qty"]))
        if o["notes"] != n["notes"]:
            note_changes.append((k, o["notes"], n["notes"]))

    print(f"  QTY changed:                       {len(qty_changes)}")
    print(f"  Notes changed:                     {len(note_changes)}")

    # Group by PV for readability
    from collections import defaultdict
    by_pv_added = defaultdict(list)
    by_pv_removed = defaultdict(list)
    by_pv_qty = defaultdict(list)

    for (pv, ms) in added:
        by_pv_added[pv].append((ms, new_m[(pv, ms)]["qty"]))
    for (pv, ms) in removed:
        by_pv_removed[pv].append((ms, old_m[(pv, ms)]["qty"]))
    for (pv, ms), oq, nq in qty_changes:
        by_pv_qty[pv].append((ms, oq, nq))

    if added:
        print(f"\n--- NEW BOM lines  ({len(added)} total, grouped by PV) ---")
        for pv in sorted(by_pv_added):
            print(f"  {pv}:")
            for ms, q in by_pv_added[pv]:
                print(f"    + {ms:48s}  qty={q}")

    if removed:
        print(f"\n--- REMOVED BOM lines  ({len(removed)} total, grouped by PV) ---")
        for pv in sorted(by_pv_removed):
            print(f"  {pv}:")
            for ms, q in by_pv_removed[pv]:
                print(f"    - {ms:48s}  was-qty={q}")

    if qty_changes:
        print(f"\n--- QTY changes  ({len(qty_changes)} total, grouped by PV) ---")
        for pv in sorted(by_pv_qty):
            print(f"  {pv}:")
            for ms, oq, nq in by_pv_qty[pv]:
                print(f"    ~ {ms:48s}  {oq} -> {nq}")


def main():
    print(f"OLD: {OLD_XLSX}")
    print(f"NEW: {NEW_XLSX}")
    print(f"BOM sheets: {OLD_BOM_SHEET} (v5)  vs  {NEW_BOM_SHEET} (v6)\n")
    if not OLD_XLSX.is_file():
        print("!! Old file not found"); return
    if not NEW_XLSX.is_file():
        print("!! New file not found"); return
    compare_materials()
    compare_bom()


if __name__ == "__main__":
    main()
