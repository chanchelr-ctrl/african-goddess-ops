"""One-shot comparison: spreadsheet rows vs DB rows.

Reads the same .xlsx files migrate_bom_xlsx imports from, and for every row
with a SKU compares: pack_size, pack_cost, unit_cost, sohand stock, reorder.
Reports any mismatch.
"""

from __future__ import annotations

import os
import sys
import django
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

import re

import openpyxl
from inventory.models import RawMaterial


def _tok(v) -> str:
    if v in (None, ""):
        return ""
    return re.sub(r"[^a-zA-Z0-9]+", "_", str(v).strip()).strip("_")


def make_sku(base_sku: str, size_val, colour_val=None) -> str:
    sku = base_sku
    for t in (_tok(size_val), _tok(colour_val)):
        if t:
            sku = f"{sku}_{t}"
    return sku[:64]

DATA_DIR = Path(__file__).resolve().parent.parent.parent / \
    "Inventory & Sales Management" / "01_Client_Analysis" / "client_data"

IGNORED_SHEETS = {"AGCCLRDBeads", "MASTERBeadSheet",
                  "AGCFindingsMadeInHouse", "BeadPlotting"}


def to_dec(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def find_header_row(rows):
    for i, row in enumerate(rows[:5]):
        non_empty = [c for c in row if c not in (None, "")]
        string_like = [c for c in non_empty if isinstance(c, str)]
        if non_empty and len(string_like) >= max(2, len(non_empty) - 1):
            return i
    return 0


def main():
    files = sorted(DATA_DIR.glob("*.xlsx"))
    rows_seen = 0
    rows_matched = 0
    missing = []
    mismatches = []
    last_xlsx_for_sku = {}
    all_rows_for_sku = {}  # sku -> list[(file, sheet, rownum, xlsx)]

    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in IGNORED_SHEETS:
                continue
            ws = wb[sheet_name]
            grid = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(grid) < 3:
                continue
            hdr_idx = find_header_row(grid)
            header = [str(c).strip() if c is not None else "" for c in grid[hdr_idx]]
            col = {n: i for i, n in enumerate(header) if n}

            sku_idx = col.get("CURRENT ITEM ID CODE")
            if sku_idx is None:
                continue

            for r in range(hdr_idx + 1, len(grid)):
                row = grid[r]
                if sku_idx >= len(row) or not row[sku_idx]:
                    continue
                base_sku = str(row[sku_idx]).strip()
                if not base_sku:
                    continue

                def cell(name):
                    i = col.get(name)
                    if i is None or i >= len(row):
                        return None
                    return row[i]

                sku = make_sku(base_sku, cell("ITEM SIZE"),
                               cell("COLOUR:") or cell("COLOUR"))

                xlsx = {
                    "pack_size": to_dec(cell("PACK SIZE")),
                    "pack_cost": to_dec(cell("TEMU PACK COST")),
                    "unit_cost": to_dec(cell("FINAL PER UNIT PRICE")),
                    "sohand": to_dec(cell("SOHAND UNIT")),
                    "reorder": to_dec(cell("UNIT QTY FOR BASE STOCK LEVEL")),
                }
                rows_seen += 1
                last_xlsx_for_sku[sku] = (f.name, sheet_name, r + 1, xlsx)
                all_rows_for_sku.setdefault(sku, []).append(
                    (f.name, sheet_name, r + 1, xlsx))

    print(f"Spreadsheet rows with a SKU: {rows_seen}")
    print(f"Unique SKUs (last-row-wins): {len(last_xlsx_for_sku)}")
    print()

    for sku, (fname, sheet, rownum, xlsx) in last_xlsx_for_sku.items():
        try:
            m = RawMaterial.objects.get(sku=sku)
        except RawMaterial.DoesNotExist:
            missing.append((sku, fname, sheet, rownum))
            continue
        rows_matched += 1

        diffs = []

        def cmp(label, x, db, places=4):
            if x is None and (db is None or db == 0):
                return
            if x is None:
                return
            xq = Decimal(x).quantize(Decimal(10) ** -places)
            dq = Decimal(db).quantize(Decimal(10) ** -places)
            if xq != dq:
                diffs.append(f"{label}: xlsx={xq} db={dq}")

        cmp("pack_size", xlsx["pack_size"], m.pack_size)
        cmp("pack_cost", xlsx["pack_cost"], m.last_paid_pack_cost)
        cmp("unit_cost", xlsx["unit_cost"], m.last_paid_unit_cost)
        cmp("sohand_stock", xlsx["sohand"], m.current_stock)
        cmp("reorder_point", xlsx["reorder"], m.reorder_point)

        if diffs:
            mismatches.append((sku, fname, sheet, rownum, diffs))

    print(f"DB matched: {rows_matched}")
    print(f"Missing in DB: {len(missing)}")
    print(f"Mismatched figures: {len(mismatches)}")

    if missing:
        print("\n--- MISSING ---")
        for sku, fname, sheet, rn in missing:
            print(f"  {sku}  ({fname} / {sheet} / row {rn})")

    if mismatches:
        print("\n--- MISMATCHES (showing every spreadsheet row for that SKU) ---")
        for sku, fname, sheet, rn, diffs in mismatches:
            print(f"\n  {sku}")
            print("  app DB:")
            m = RawMaterial.objects.get(sku=sku)
            print(f"    pack_size={m.pack_size}  pack_cost={m.last_paid_pack_cost}  "
                  f"unit_cost={m.last_paid_unit_cost}  stock={m.current_stock}  "
                  f"reorder={m.reorder_point}")
            print("  spreadsheet rows:")
            for fn, sh, rownum, x in all_rows_for_sku[sku]:
                print(f"    {fn} / {sh} / row {rownum}: "
                      f"pack_size={x['pack_size']}  pack_cost={x['pack_cost']}  "
                      f"unit_cost={x['unit_cost']}  sohand={x['sohand']}  "
                      f"reorder={x['reorder']}")

    if not missing and not mismatches:
        print("\nAll spreadsheet rows reconcile to the DB.")


if __name__ == "__main__":
    main()
