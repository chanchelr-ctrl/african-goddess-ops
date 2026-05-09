"""Export master data v5 — two-sheet workbook separating operational from spec.

Sheet 1 (default) — "MaterialMaster":
  Operational. One row per RawMaterial (81 rows × 22 cols), Stock on Hand
  as the last column. Same shape as v3. Refreshed from the live DB; changes
  daily as stock moves and costs update.

Sheet 2 — "ProductSpec":
  Design-controlled reference data. One row per (ProductVariant, Material)
  spec line — i.e. one row per BOM entry — but containing ONLY spec data:
  product / variant / brand / retail price / material identity (sku, name,
  size, colour, shape) / BOM quantity / notes. No stock, no cost, no
  supplier — those live on Sheet 1 and can be looked up via Material SKU.

Output: ../Inventory & Sales Management/01_Client_Analysis/client_data_v5/MasterData_v5.xlsx
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
import django
django.setup()

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inventory.models import BomLine, RawMaterial

OUT_DIR = Path(__file__).resolve().parent.parent.parent / \
    "Inventory & Sales Management" / "01_Client_Analysis" / "client_data_v5"
OUT_FILE = OUT_DIR / "MasterData_v5.xlsx"


# ---------- helpers ----------------------------------------------------------


def supplier_id_from_sku(sku: str) -> str:
    return sku.split("_", 1)[0]


def style_header_row(ws):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")


def autosize(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), max_width)


def fmt_columns(ws, headers, cols, fmt):
    for label in cols:
        if label not in headers:
            continue
        c = headers.index(label) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = fmt


# ---------- Sheet 1: MaterialMaster (v3 layout) -----------------------------


SHEET1_HEADERS = [
    "SKU",
    "Supplier ID (Current)",
    "Alternative ID",
    "Internal ID",
    "Name",
    "Description",
    "Item Size",
    "Colour",
    "Finish",
    "Shape",
    "Sub-brand",
    "Unit",
    "Pack Size",
    "Pack Cost (ZAR)",
    "Import Duties / Pack",
    "Unit Cost (ZAR)",
    "Supplier",
    "Reorder Point",
    "Reorder Quantity",
    "Active?",
    "Notes",
    "Stock on Hand",
    "ZAR Value of Stock on Hand",  # = Stock * Unit Cost (live formula)
]


CURRENCY_FMT = '"R "#,##0.00'   # ZAR with 2 decimals
WHOLE_NUMBER_FMT = "#,##0"      # whole numbers with thousands separator


def build_material_master_sheet(wb):
    ws = wb.active
    ws.title = "MaterialMaster"
    ws.append(SHEET1_HEADERS)

    unit_cost_col = SHEET1_HEADERS.index("Unit Cost (ZAR)") + 1
    stock_col = SHEET1_HEADERS.index("Stock on Hand") + 1
    value_col = SHEET1_HEADERS.index("ZAR Value of Stock on Hand") + 1
    unit_cost_letter = get_column_letter(unit_cost_col)
    stock_letter = get_column_letter(stock_col)

    for m in RawMaterial.objects.select_related("preferred_supplier").order_by("sku"):
        ws.append([
            m.sku,
            supplier_id_from_sku(m.sku),
            m.alternative_id_code or "",
            m.internal_id_code or "",
            m.name,
            m.description or "",
            m.item_size or "",
            m.colour or "",
            m.finish or "",
            m.shape or "",
            m.sub_brand or "",
            m.get_unit_display(),
            float(m.pack_size),
            float(m.last_paid_pack_cost),
            float(m.import_duties_per_pack),
            float(m.last_paid_unit_cost),
            m.preferred_supplier.name if m.preferred_supplier else "",
            float(m.reorder_point),
            float(m.reorder_quantity),
            "Yes" if m.is_active else "No",
            m.notes or "",
            float(m.current_stock),
            None,  # placeholder; formula written below
        ])

    # Live formula for ZAR Value of Stock on Hand = Stock * Unit Cost
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=value_col).value = f"={stock_letter}{r}*{unit_cost_letter}{r}"

    style_header_row(ws)
    fmt_columns(ws, SHEET1_HEADERS,
                ["Pack Size", "Reorder Point", "Reorder Quantity", "Stock on Hand"],
                WHOLE_NUMBER_FMT)
    fmt_columns(ws, SHEET1_HEADERS,
                ["Pack Cost (ZAR)", "Import Duties / Pack", "Unit Cost (ZAR)",
                 "ZAR Value of Stock on Hand"],
                CURRENCY_FMT)
    ws.freeze_panes = "B2"
    autosize(ws)
    print(f"  Sheet 'MaterialMaster': {ws.max_row - 1} rows x {ws.max_column} cols")


# ---------- Sheet 2: ProductSpec (long-form, spec-only) ---------------------


SHEET2_HEADERS = [
    "PV SKU",
    "Product Code",
    "Product Name",
    "Variant Code",
    "Variant Name",
    "Brand",
    "Retail Price (ZAR)",
    "Material SKU",
    "Material Name",
    "Item Size",
    "Colour",
    "Shape",
    "BOM Quantity",
    "Notes",
]


def build_product_spec_sheet(wb):
    ws = wb.create_sheet("ProductSpec")
    ws.append(SHEET2_HEADERS)

    qs = (BomLine.objects
          .select_related("raw_material",
                          "product_variant",
                          "product_variant__product",
                          "product_variant__product__brand",
                          "product_variant__variant")
          .order_by("product_variant__product__brand__code",
                    "product_variant__variant__code",
                    "product_variant__product__code",
                    "raw_material__sku"))

    for line in qs:
        m = line.raw_material
        pv = line.product_variant
        ws.append([
            pv.sku,
            pv.product.code,
            pv.product.name,
            pv.variant.code,
            pv.variant.name,
            pv.product.brand.name,
            float(pv.effective_retail_price_zar),
            m.sku,
            m.name,
            m.item_size or "",
            m.colour or "",
            m.shape or "",
            float(line.quantity),
            line.notes or "",
        ])

    style_header_row(ws)
    fmt_columns(ws, SHEET2_HEADERS, ["BOM Quantity"], WHOLE_NUMBER_FMT)
    fmt_columns(ws, SHEET2_HEADERS, ["Retail Price (ZAR)"], CURRENCY_FMT)
    ws.freeze_panes = "B2"
    autosize(ws)
    print(f"  Sheet 'ProductSpec': {ws.max_row - 1} rows x {ws.max_column} cols")


# ---------- main ------------------------------------------------------------


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    print(f"Building {OUT_FILE.name}")
    build_material_master_sheet(wb)
    build_product_spec_sheet(wb)
    wb.active = 0
    wb.save(OUT_FILE)
    print(f"\nSaved -> {OUT_FILE}")


if __name__ == "__main__":
    main()
