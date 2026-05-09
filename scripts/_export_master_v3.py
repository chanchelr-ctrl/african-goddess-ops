"""Export every RawMaterial as a flat .xlsx — one row per material, with
current stock as the last column.

Output: ../Inventory & Sales Management/01_Client_Analysis/client_data_v3/MasterData_v3.xlsx
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
import django
django.setup()

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inventory.models import RawMaterial

OUT_DIR = Path(__file__).resolve().parent.parent.parent / \
    "Inventory & Sales Management" / "01_Client_Analysis" / "client_data_v3"
OUT_FILE = OUT_DIR / "MasterData_v3.xlsx"


COLUMNS = [
    ("SKU",                      lambda m: m.sku),
    ("Supplier ID (Current)",    lambda m: _supplier_id_from_sku(m.sku)),
    ("Alternative ID",           lambda m: m.alternative_id_code or ""),
    ("Internal ID",              lambda m: m.internal_id_code or ""),
    ("Name",                     lambda m: m.name),
    ("Description",              lambda m: m.description or ""),
    ("Item Size",                lambda m: m.item_size or ""),
    ("Colour",                   lambda m: m.colour or ""),
    ("Finish",                   lambda m: m.finish or ""),
    ("Shape",                    lambda m: m.shape or ""),
    ("Sub-brand",                lambda m: m.sub_brand or ""),
    ("Unit",                     lambda m: m.get_unit_display()),
    ("Pack Size",                lambda m: float(m.pack_size)),
    ("Pack Cost (ZAR)",          lambda m: float(m.last_paid_pack_cost)),
    ("Import Duties / Pack",     lambda m: float(m.import_duties_per_pack)),
    ("Unit Cost (ZAR)",          lambda m: float(m.last_paid_unit_cost)),
    ("Supplier",                 lambda m: m.preferred_supplier.name if m.preferred_supplier else ""),
    ("Reorder Point",            lambda m: float(m.reorder_point)),
    ("Reorder Quantity",         lambda m: float(m.reorder_quantity)),
    ("Active?",                  lambda m: "Yes" if m.is_active else "No"),
    ("Notes",                    lambda m: m.notes or ""),
    ("Stock on Hand",            lambda m: float(m.current_stock)),  # MUST stay last
]


def _supplier_id_from_sku(sku: str) -> str:
    """The disambiguated SKU is `{supplierId}_{size}_{colour}` — strip the
    suffix to recover the original supplier ID code."""
    return sku.split("_", 1)[0]


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MasterData"

    headers = [c[0] for c in COLUMNS]
    ws.append(headers)

    # Header style
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows — sorted by SKU for stable output
    materials = RawMaterial.objects.select_related("preferred_supplier").order_by("sku")
    for m in materials:
        ws.append([fn(m) for _, fn in COLUMNS])

    # Number format on the numeric columns
    money_cols = ["Pack Cost (ZAR)", "Import Duties / Pack", "Unit Cost (ZAR)"]
    qty_cols = ["Pack Size", "Reorder Point", "Reorder Quantity", "Stock on Hand"]
    for label in money_cols:
        col = headers.index(label) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "#,##0.0000"
    for label in qty_cols:
        col = headers.index(label) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = "#,##0.0000"

    # Auto-size columns (cap to keep file readable)
    for col_idx, label in enumerate(headers, start=1):
        max_len = len(label)
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Freeze the header row + first column (SKU) for easy scrolling
    ws.freeze_panes = "B2"

    wb.save(OUT_FILE)
    print(f"Wrote {ws.max_row - 1} rows x {len(headers)} cols  ->  {OUT_FILE}")


if __name__ == "__main__":
    main()
