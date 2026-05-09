"""Inventory every BEAD TOTALS row in the v2 spreadsheets and report:
  - which sheet
  - which row
  - which product columns it sits across
  - which data rows belong to its variant section above it
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent.parent / \
    "Inventory & Sales Management" / "01_Client_Analysis" / "client_data_v2"

# Same ignore-list as the importer
IGNORED_SHEETS = {"AGCCLRDBeads", "MASTERBeadSheet",
                  "AGCFindingsMadeInHouse", "BeadPlotting"}

VARIANT_RE = re.compile(r"^\s*(SB[RA]\d{2})\b", re.IGNORECASE)
PRODUCT_CODE_RE = re.compile(r"^SB[RA]00[A-Z]+$")


def find_header_row(rows):
    """Mirror the importer's header detector."""
    for i, row in enumerate(rows[:5]):
        non_empty = [c for c in row if c not in (None, "")]
        string_like = [c for c in non_empty if isinstance(c, str)]
        if non_empty and len(string_like) >= max(2, len(non_empty) - 1):
            return i
    return 0


def is_bead_totals_row(row, descr_idx) -> bool:
    if descr_idx is None or descr_idx >= len(row):
        return False
    v = row[descr_idx]
    if not v:
        return False
    return "BEAD TOTALS" in str(v).upper()


def main():
    files = sorted(DATA_DIR.glob("*.xlsx"))
    if not files:
        print(f"No xlsx in {DATA_DIR}")
        sys.exit(1)

    for f in files:
        print(f"\n=== {f.name} ===")
        wb = openpyxl.load_workbook(f, data_only=True, read_only=False)
        for sheet_name in wb.sheetnames:
            if sheet_name in IGNORED_SHEETS:
                continue
            ws = wb[sheet_name]
            print(f"\n  Sheet: {sheet_name}  (max_row={ws.max_row})")

            rows = list(ws.iter_rows(values_only=True))
            hdr_idx = find_header_row(rows)
            header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
            col = {n: i for i, n in enumerate(header) if n}

            cc_idx = col.get("COLOUR COMBINATION")
            sku_idx = col.get("CURRENT ITEM ID CODE")
            descr_idx = col.get("DESCRIPTION")

            # Detect product columns (those whose header looks like a product code)
            product_cols = [(i, h) for i, h in enumerate(header)
                            if h and PRODUCT_CODE_RE.match(h)]
            print(f"    Product columns ({len(product_cols)}):")
            for i, h in product_cols:
                print(f"      col {get_column_letter(i + 1)} ({i + 1}): {h}")

            # Walk rows, mark variant boundaries + bead-totals rows
            current_variant = None
            section_start_xlsx = None
            for r_idx in range(hdr_idx + 1, len(rows)):
                row = rows[r_idx]
                xlsx_row = r_idx + 1
                cc_val = row[cc_idx] if (cc_idx is not None and cc_idx < len(row)) else None
                cc_str = str(cc_val).strip() if cc_val else ""
                m = VARIANT_RE.match(cc_str) if cc_str else None
                if m:
                    current_variant = m.group(1).upper()
                    section_start_xlsx = xlsx_row  # this row IS the first data row
                if is_bead_totals_row(row, descr_idx):
                    section_end_xlsx = xlsx_row - 1  # data rows up to the row before TOTALS
                    print(f"    BEAD TOTALS row {xlsx_row}  variant={current_variant}  "
                          f"data rows {section_start_xlsx}..{section_end_xlsx} "
                          f"({section_end_xlsx - section_start_xlsx + 1 if section_start_xlsx else 0} rows)")


if __name__ == "__main__":
    main()
