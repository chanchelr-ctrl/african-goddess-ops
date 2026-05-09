"""Export a two-sheet master data workbook for review.

Sheet 1 (default) — "BOM":
  Long-form BOM join. One row per BomLine, with material attributes,
  product-variant attributes, BOM quantity, and material stock-on-hand
  denormalised onto every row.

Sheet 2 — "MaterialRollup":
  One row per RawMaterial (matches v3 shape) with extra rolled-up usage
  columns: how many product variants use it, which product types,
  which variants, total quantity per "one of each PV", and the single
  largest BOM use. Stock on Hand is the last column.

Output: ../Inventory & Sales Management/01_Client_Analysis/client_data_v4/MasterData_v4.xlsx
"""

from __future__ import annotations

import os
import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
import django
django.setup()

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from inventory.models import BomLine, ProductVariant, RawMaterial

OUT_DIR = Path(__file__).resolve().parent.parent.parent / \
    "Inventory & Sales Management" / "01_Client_Analysis" / "client_data_v4"
OUT_FILE = OUT_DIR / "MasterData_v4.xlsx"

ZERO = Decimal("0")

# ---------- helpers ----------------------------------------------------------


def supplier_id_from_sku(sku: str) -> str:
    """Recover the original supplier listing ID from the disambiguated SKU."""
    return sku.split("_", 1)[0]


def short_product_type(product_code: str) -> str:
    """SBR00EARR -> EARR, SBA00MULTI -> MULTI, SBA00FBA -> FBA."""
    if len(product_code) > 5 and product_code[:5] in ("SBR00", "SBA00"):
        return product_code[5:]
    return product_code


def style_header_row(ws):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for col_idx in range(1, ws.max_column + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")


def autosize(ws, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), max_width)


def fmt_numeric_columns(ws, headers, cols, fmt):
    for label in cols:
        if label not in headers:
            continue
        c = headers.index(label) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = fmt


# ---------- Sheet 1: BOM (long-form join) -----------------------------------


SHEET1_HEADERS = [
    "Material SKU",
    "Supplier ID (Current)",
    "Internal ID",
    "Material Name",
    "Description",
    "Item Size",
    "Colour",
    "Finish",
    "Shape",
    "Sub-brand",
    "Unit",
    "Pack Size",
    "Pack Cost (ZAR)",
    "Unit Cost (ZAR)",
    "Supplier",
    "Reorder Point",
    "PV SKU",
    "Product Type",
    "Variant",
    "Brand",
    "Retail Price (ZAR)",
    "BOM Quantity",
    "Cost Contribution (ZAR)",
    "Stock on Hand",  # MUST stay last for consistency with v3
]


def build_bom_sheet(wb):
    ws = wb.active
    ws.title = "BOM"
    ws.append(SHEET1_HEADERS)

    qs = (BomLine.objects
          .select_related("raw_material",
                          "raw_material__preferred_supplier",
                          "product_variant",
                          "product_variant__product",
                          "product_variant__product__brand",
                          "product_variant__variant")
          .order_by("raw_material__sku", "product_variant__sku"))

    for line in qs:
        m = line.raw_material
        pv = line.product_variant
        cost_contrib = (Decimal(line.quantity) * m.last_paid_unit_cost).quantize(Decimal("0.0001"))
        retail = pv.effective_retail_price_zar
        ws.append([
            m.sku,
            supplier_id_from_sku(m.sku),
            m.internal_id_code or "",
            m.name,
            m.description or "",
            m.item_size or "",
            m.colour or "",
            m.finish or "",
            m.shape or "",
            m.sub_brand or "",
            m.get_unit_display(),
            float(m.pack_size),
            float(m.last_paid_pack_cost),
            float(m.last_paid_unit_cost),
            m.preferred_supplier.name if m.preferred_supplier else "",
            float(m.reorder_point),
            pv.sku,
            short_product_type(pv.product.code),
            pv.variant.code,
            pv.product.brand.name,
            float(retail),
            float(line.quantity),
            float(cost_contrib),
            float(m.current_stock),
        ])

    style_header_row(ws)
    fmt_numeric_columns(ws, SHEET1_HEADERS,
                        ["Pack Size", "Reorder Point", "BOM Quantity", "Stock on Hand"],
                        "#,##0.0000")
    fmt_numeric_columns(ws, SHEET1_HEADERS,
                        ["Pack Cost (ZAR)", "Unit Cost (ZAR)",
                         "Retail Price (ZAR)", "Cost Contribution (ZAR)"],
                        "#,##0.0000")
    ws.freeze_panes = "B2"
    autosize(ws)
    print(f"  Sheet 'BOM': {ws.max_row - 1} rows x {ws.max_column} cols")


# ---------- Sheet 2: MaterialRollup -----------------------------------------


SHEET2_HEADERS = [
    "Material SKU",
    "Supplier ID (Current)",
    "Internal ID",
    "Material Name",
    "Description",
    "Item Size",
    "Colour",
    "Finish",
    "Shape",
    "Sub-brand",
    "Unit",
    "Pack Size",
    "Pack Cost (ZAR)",
    "Unit Cost (ZAR)",
    "Supplier",
    "Reorder Point",
    "Used in (# PVs)",
    "Used in product types",
    "Used in variants",
    "Qty per one-of-each-PV",
    "Largest single use",
    "Stock on Hand",  # last
]


def build_rollup_sheet(wb):
    ws = wb.create_sheet("MaterialRollup")
    ws.append(SHEET2_HEADERS)

    # Pre-compute usage rollups: material_id -> { pv_count, types, variants, total_qty, largest }
    rollup: dict[int, dict] = defaultdict(lambda: {
        "pv_count": 0,
        "types": set(),
        "variants": set(),
        "total_qty": Decimal("0"),
        "largest_qty": Decimal("0"),
        "largest_pv_sku": "",
    })

    bom_qs = (BomLine.objects
              .select_related("raw_material",
                              "product_variant",
                              "product_variant__product",
                              "product_variant__variant"))
    for line in bom_qs:
        r = rollup[line.raw_material_id]
        r["pv_count"] += 1
        r["types"].add(short_product_type(line.product_variant.product.code))
        r["variants"].add(line.product_variant.variant.code)
        r["total_qty"] += Decimal(line.quantity)
        if line.quantity > r["largest_qty"]:
            r["largest_qty"] = Decimal(line.quantity)
            r["largest_pv_sku"] = line.product_variant.sku

    # Stable type/variant ordering
    type_order = ["EARR", "DNL", "DBL", "DWC", "DANK", "BBS",
                  "NL", "BK", "SN", "SL", "WC", "MULTI", "FBA"]
    type_rank = {t: i for i, t in enumerate(type_order)}

    for m in RawMaterial.objects.select_related("preferred_supplier").order_by("sku"):
        r = rollup.get(m.id, {
            "pv_count": 0, "types": set(), "variants": set(),
            "total_qty": ZERO, "largest_qty": ZERO, "largest_pv_sku": "",
        })
        types_str = ", ".join(sorted(r["types"], key=lambda t: type_rank.get(t, 999)))
        variants_str = ", ".join(sorted(r["variants"]))
        if r["largest_qty"] > 0:
            largest_str = f"{float(r['largest_qty']):g} in {r['largest_pv_sku']}"
        else:
            largest_str = ""

        ws.append([
            m.sku,
            supplier_id_from_sku(m.sku),
            m.internal_id_code or "",
            m.name,
            m.description or "",
            m.item_size or "",
            m.colour or "",
            m.finish or "",
            m.shape or "",
            m.sub_brand or "",
            m.get_unit_display(),
            float(m.pack_size),
            float(m.last_paid_pack_cost),
            float(m.last_paid_unit_cost),
            m.preferred_supplier.name if m.preferred_supplier else "",
            float(m.reorder_point),
            r["pv_count"],
            types_str,
            variants_str,
            float(r["total_qty"]),
            largest_str,
            float(m.current_stock),
        ])

    style_header_row(ws)
    fmt_numeric_columns(ws, SHEET2_HEADERS,
                        ["Pack Size", "Reorder Point", "Qty per one-of-each-PV", "Stock on Hand"],
                        "#,##0.0000")
    fmt_numeric_columns(ws, SHEET2_HEADERS,
                        ["Pack Cost (ZAR)", "Unit Cost (ZAR)"],
                        "#,##0.0000")
    ws.freeze_panes = "B2"
    autosize(ws)
    print(f"  Sheet 'MaterialRollup': {ws.max_row - 1} rows x {ws.max_column} cols")


# ---------- main ------------------------------------------------------------


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    print(f"Building {OUT_FILE.name}")
    build_bom_sheet(wb)
    build_rollup_sheet(wb)
    wb.active = 0  # 'BOM' sheet opens by default
    wb.save(OUT_FILE)
    print(f"\nSaved -> {OUT_FILE}")


if __name__ == "__main__":
    main()
