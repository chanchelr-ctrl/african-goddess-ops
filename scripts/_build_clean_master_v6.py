"""
Build MasterData_v6.xlsx — clean, importable canonical workbook derived
from the client's v6-1 submission.

Inputs:
  client_data_v5/MasterData_v5.xlsx   (last clean canonical, used for the
                                       PV-level metadata lookup: brand, product
                                       code, product name, variant code,
                                       variant name)
  client_data_v6/MasterData_v6-1.xlsx (client's edited workbook — TUSHY sheet
                                       is the new BOM source-of-truth)

Output:
  client_data_v6/MasterData_v6.xlsx with four sheets:
    1. MaterialMaster  — copied verbatim from v6-1 (unchanged from v5)
    2. ProductSpec     — built from TUSHY, but in the canonical 14-col schema
                         that import_master expects. PV-level metadata
                         (brand, product / variant names) re-derived from v5
                         so the TUSHY brand mislabel is corrected.
    3. ProductPricing  — NEW. One row per PV, deduped from TUSHY. Captures
                         Retail + Wholesale price. Importer ignores this for
                         now; preserved for the future pricing feature.
    4. ChangeLog       — empty (header only), reserved for the round-trip
                         audit log.

Run (from project root):
    .\\.venv\\Scripts\\python.exe scripts\\_build_clean_master_v6.py
"""
from __future__ import annotations

from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CLIENT_DIR = Path(r"C:\Users\chanr\Desktop\Inventory & Sales Management\01_Client_Analysis")
V5_XLSX  = CLIENT_DIR / "client_data_v5" / "MasterData_v5.xlsx"
RAW_XLSX = CLIENT_DIR / "client_data_v6" / "MasterData_v6-1.xlsx"
OUT_XLSX = CLIENT_DIR / "client_data_v6" / "MasterData_v6.xlsx"

PRODUCT_SPEC_HEADERS = [
    "PV SKU", "Product Code", "Product Name", "Variant Code", "Variant Name",
    "Brand", "Retail Price (ZAR)", "Material SKU", "Material Name",
    "Item Size", "Colour", "Shape", "BOM Quantity", "Notes",
]
PRODUCT_PRICING_HEADERS = [
    "PV SKU", "Brand", "Variant Name", "Retail Price (ZAR)", "Wholesale Price (ZAR)",
]
CHANGELOG_HEADERS = [
    "Timestamp", "User", "Action", "Model", "SKU/Code",
    "Object PK", "Field", "Old Value", "New Value", "Note",
]


def _norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _num(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


def read_sheet(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        return None, []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [_norm(c) for c in rows[0]]
    data = []
    for r in rows[1:]:
        if all(c is None or _norm(c) == "" for c in r):
            continue
        data.append(r)
    return header, data


def to_dict(header, row):
    return {header[i]: row[i] if i < len(row) else None for i in range(len(header))}


# ---------------------------------------------------------------------------
# Build PV-level metadata lookup from v5 (the canonical truth for brand, etc.)
# ---------------------------------------------------------------------------

def build_pv_lookup_from_v5():
    """Map: PV SKU -> {Product Code, Product Name, Variant Code, Variant Name, Brand}"""
    header, rows = read_sheet(V5_XLSX, "ProductSpec")
    lookup = {}
    for r in rows:
        d = to_dict(header, r)
        pv = _norm(d.get("PV SKU"))
        if not pv or pv in lookup:
            continue
        lookup[pv] = {
            "Product Code":  _norm(d.get("Product Code")),
            "Product Name":  _norm(d.get("Product Name")),
            "Variant Code":  _norm(d.get("Variant Code")),
            "Variant Name":  _norm(d.get("Variant Name")),
            "Brand":         _norm(d.get("Brand")),
        }
    return lookup


# ---------------------------------------------------------------------------
# Build ProductSpec rows from TUSHY
# ---------------------------------------------------------------------------

def build_product_spec_rows(pv_lookup):
    header, rows = read_sheet(RAW_XLSX, "TUSHY")

    out_rows = []
    unknown_pvs = set()

    for r in rows:
        d = to_dict(header, r)
        pv = _norm(d.get("PV SKU"))
        ms = _norm(d.get("Material SKU"))
        if not pv or not ms:
            continue

        meta = pv_lookup.get(pv)
        if not meta:
            # PV exists in TUSHY but not in v5 — flag and still include
            unknown_pvs.add(pv)
            meta = {
                "Product Code":  _norm(d.get("Product Code")),
                "Product Name":  _norm(d.get("Product Name")),
                "Variant Code":  _norm(d.get("Variant Code")),
                "Variant Name":  _norm(d.get("Variant Name")),
                "Brand":         _norm(d.get("Brand")),
            }

        # Retail price: keep TUSHY's value (this is the genuinely new data)
        retail = d.get("Retail Price (ZAR)")
        if retail in (None, ""):
            retail = None
        else:
            try:
                retail = float(retail)
            except (TypeError, ValueError):
                retail = None

        qty = _num(d.get("BOM Quantity"))
        try:
            qty_val = float(qty) if qty is not None else 0
        except (TypeError, ValueError):
            qty_val = 0

        out_rows.append([
            pv,
            meta["Product Code"],
            meta["Product Name"],
            meta["Variant Code"],
            meta["Variant Name"],
            meta["Brand"],
            retail,
            ms,
            _norm(d.get("Material Name")),
            _norm(d.get("Item Size")),
            _norm(d.get("Colour")),
            _norm(d.get("Shape")),
            qty_val,
            _norm(d.get("Notes")),
        ])

    return out_rows, sorted(unknown_pvs)


# ---------------------------------------------------------------------------
# Build ProductPricing — one row per PV, deduped from TUSHY
# ---------------------------------------------------------------------------

def build_pricing_rows(pv_lookup):
    header, rows = read_sheet(RAW_XLSX, "TUSHY")

    from collections import Counter
    seen = {}            # pv -> (retail, wholesale)
    inconsistencies = [] # list of (pv, counter of (retail, wholesale))

    pv_price_counts = defaultdict(Counter)
    for r in rows:
        d = to_dict(header, r)
        pv = _norm(d.get("PV SKU"))
        if not pv:
            continue
        retail = d.get("Retail Price (ZAR)")
        wholesale = d.get("Wholesale Price (ZAR)")
        try:
            retail = float(retail) if retail not in (None, "") else None
        except (TypeError, ValueError):
            retail = None
        try:
            wholesale = float(wholesale) if wholesale not in (None, "") else None
        except (TypeError, ValueError):
            wholesale = None
        pv_price_counts[pv][(retail, wholesale)] += 1

    for pv, counter in pv_price_counts.items():
        # Drop the (None, None) entry from candidates if there's any real price
        real = {p: c for p, c in counter.items() if p != (None, None)}
        candidates = real if real else dict(counter)
        if len(candidates) > 1:
            inconsistencies.append((pv, counter))
        # Pick the most-frequent price (mode). Tie-break: higher retail wins —
        # placeholders tend to be lower than real prices.
        chosen = max(candidates.items(), key=lambda kv: (kv[1], (kv[0][0] or 0)))[0]
        seen[pv] = chosen

    out_rows = []
    for pv in sorted(seen):
        retail, wholesale = seen[pv]
        meta = pv_lookup.get(pv, {})
        out_rows.append([
            pv,
            meta.get("Brand", ""),
            meta.get("Variant Name", ""),
            retail,
            wholesale,
        ])

    return out_rows, inconsistencies


# ---------------------------------------------------------------------------
# Build MaterialMaster — verbatim copy of v6-1
# ---------------------------------------------------------------------------

def copy_material_master():
    header, rows = read_sheet(RAW_XLSX, "MaterialMaster")
    return header, [list(r) for r in rows]


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="C89759")  # warm gold (brand colour)
HEADER_FONT = Font(bold=True, color="3D2013")
WHOLE_NUMBER_FMT = "#,##0"
CURRENCY_FMT = "R #,##0.00"


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def autosize(ws, max_width=42):
    for col in ws.columns:
        # column iterator can yield empty in read_only; this is write mode
        col_list = list(col)
        if not col_list:
            continue
        letter = get_column_letter(col_list[0].column)
        width = min(max_width, max((len(str(c.value)) for c in col_list if c.value is not None), default=10) + 2)
        ws.column_dimensions[letter].width = width


def format_columns(ws, headers, column_names, number_format):
    for col_name in column_names:
        if col_name not in headers:
            continue
        col_idx = headers.index(col_name) + 1
        letter = get_column_letter(col_idx)
        for cell in ws[letter][1:]:
            cell.number_format = number_format


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print(f"v5 reference: {V5_XLSX}")
    print(f"v6 raw:       {RAW_XLSX}")
    print(f"v6 clean out: {OUT_XLSX}\n")

    pv_lookup = build_pv_lookup_from_v5()
    print(f"PV metadata loaded from v5: {len(pv_lookup)} PVs")

    # ProductSpec
    ps_rows, unknown_pvs = build_product_spec_rows(pv_lookup)
    print(f"ProductSpec rows built from TUSHY: {len(ps_rows)}")
    if unknown_pvs:
        print(f"  ! TUSHY has {len(unknown_pvs)} PV(s) not in v5: {unknown_pvs}")

    # ProductPricing
    pricing_rows, inconsistencies = build_pricing_rows(pv_lookup)
    print(f"ProductPricing rows (deduped): {len(pricing_rows)}")
    if inconsistencies:
        print(f"  ! Pricing inconsistencies for {len(inconsistencies)} PV(s)"
              f" — picked most-common price (placeholder rows ignored):")
        for pv, counter in inconsistencies:
            picked = pricing_rows_for_pv = next((r for r in pricing_rows if r[0] == pv), None)
            print(f"    {pv}:")
            for price, count in counter.most_common():
                marker = "  <-- picked" if picked and (price[0] == picked[3] and price[1] == picked[4]) else ""
                print(f"        retail={price[0]} wholesale={price[1]}   rows={count}{marker}")

    # MaterialMaster
    mm_header, mm_rows = copy_material_master()
    print(f"MaterialMaster rows copied: {len(mm_rows)}")

    # ----- Build workbook -----
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # MaterialMaster
    ws = wb.create_sheet("MaterialMaster")
    ws.append(mm_header)
    for r in mm_rows:
        ws.append(r)
    style_header(ws)
    ws.freeze_panes = "B2"
    format_columns(ws, mm_header,
                   ["Pack Size", "Reorder Point", "Reorder Quantity", "Stock on Hand"],
                   WHOLE_NUMBER_FMT)
    format_columns(ws, mm_header,
                   ["Pack Cost (ZAR)", "Import Duties / Pack", "Unit Cost (ZAR)"],
                   CURRENCY_FMT)
    autosize(ws)

    # ProductSpec
    ws = wb.create_sheet("ProductSpec")
    ws.append(PRODUCT_SPEC_HEADERS)
    for r in ps_rows:
        ws.append(r)
    style_header(ws)
    ws.freeze_panes = "B2"
    format_columns(ws, PRODUCT_SPEC_HEADERS, ["BOM Quantity"], WHOLE_NUMBER_FMT)
    format_columns(ws, PRODUCT_SPEC_HEADERS, ["Retail Price (ZAR)"], CURRENCY_FMT)
    autosize(ws)

    # ProductPricing
    ws = wb.create_sheet("ProductPricing")
    ws.append(PRODUCT_PRICING_HEADERS)
    for r in pricing_rows:
        ws.append(r)
    style_header(ws)
    ws.freeze_panes = "B2"
    format_columns(ws, PRODUCT_PRICING_HEADERS,
                   ["Retail Price (ZAR)", "Wholesale Price (ZAR)"], CURRENCY_FMT)
    autosize(ws)

    # ChangeLog (empty, header only)
    ws = wb.create_sheet("ChangeLog")
    ws.append(CHANGELOG_HEADERS)
    style_header(ws)
    ws.freeze_panes = "B2"
    autosize(ws)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    print(f"\nWrote: {OUT_XLSX}")
    print(f"  size: {OUT_XLSX.stat().st_size:,} bytes")


if __name__ == "__main__":
    main()
