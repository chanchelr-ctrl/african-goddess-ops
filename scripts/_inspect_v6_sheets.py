"""Dump SBR06Data + summarise TUSHY scope."""
from pathlib import Path
import openpyxl

NEW = Path(r"C:\Users\chanr\Desktop\Inventory & Sales Management\01_Client_Analysis\client_data_v6\MasterData_v6-1.xlsx")
wb = openpyxl.load_workbook(NEW, data_only=True, read_only=True)

# Full dump of SBR06Data
print("=" * 70)
print("SBR06Data (full)")
print("=" * 70)
ws = wb["SBR06Data"]
for i, r in enumerate(ws.iter_rows(values_only=True), 1):
    cells = list(r)
    while cells and cells[-1] in (None, ""):
        cells.pop()
    print(f"  row {i:2d}: {cells}")

# TUSHY scope analysis
print("\n" + "=" * 70)
print("TUSHY scope")
print("=" * 70)
ws = wb["TUSHY"]
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
print(f"Headers ({len(header)}): {header}")

from collections import Counter
pv_counter = Counter()
brand_counter = Counter()
product_counter = Counter()
prices = set()
for r in rows[1:]:
    if r[0] is None:
        continue
    pv_counter[str(r[0])] += 1
    brand_counter[str(r[5]) if r[5] else ""] += 1
    product_counter[str(r[2]) if r[2] else ""] += 1
    prices.add((r[6], r[7]))

print(f"\nTotal data rows: {sum(pv_counter.values())}")
print(f"Distinct PVs:    {len(pv_counter)}")
print(f"Distinct brands: {dict(brand_counter)}")
print(f"\nProduct types in TUSHY:")
for p, c in product_counter.most_common():
    print(f"  {c:4d}  {p}")
print(f"\nUnique (retail, wholesale) pairs:")
for pr in sorted(prices, key=lambda x: (x[0] or 0)):
    print(f"  retail={pr[0]} wholesale={pr[1]}")
print(f"\nTop 30 PVs by row count:")
for pv, c in pv_counter.most_common(30):
    print(f"  {pv:14s}  {c} rows")

# Compare TUSHY's PVs vs ProductSpec's PVs
ws_ps = wb["ProductSpec"]
ps_pv = Counter()
for r in list(ws_ps.iter_rows(values_only=True))[1:]:
    if r[0]:
        ps_pv[str(r[0])] += 1

tushy_pvs = set(pv_counter)
ps_pvs = set(ps_pv)
print(f"\nProductSpec PVs:     {len(ps_pvs)}")
print(f"TUSHY PVs:           {len(tushy_pvs)}")
print(f"In TUSHY only:       {sorted(tushy_pvs - ps_pvs)}")
print(f"In ProductSpec only: {sorted(ps_pvs - tushy_pvs)}")
