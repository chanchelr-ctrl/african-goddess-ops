"""Apply correct =SUM() formulas to every BEAD TOTALS row across the v2
spreadsheets.

For each sheet, locate variant sections (delimited by a non-empty
COLOUR COMBINATION at the top and a BEAD TOTALS = row at the bottom).
For each product column (header matches ^SB[RA]00[A-Z]+$), set the
BEAD TOTALS cell to =SUM(top..bottom-1).

We also report what the expected sum would be (computed from the
already-cached values) so the user can sanity-check the result.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DATA_V2 = Path(__file__).resolve().parent.parent.parent / \
    "Inventory & Sales Management" / "01_Client_Analysis" / "client_data_v2"

PRODUCT_CODE_RE = re.compile(r"^SB[RA]00[A-Z]+$")


def find_header_row_idx(rows) -> int:
    for i, row in enumerate(rows[:5]):
        non_empty = [c for c in row if c not in (None, "")]
        string_like = [c for c in non_empty if isinstance(c, str)]
        if non_empty and len(string_like) >= max(2, len(non_empty) - 1):
            return i
    return 0


def detect_columns(header):
    cc_idx = header.index("COLOUR COMBINATION") if "COLOUR COMBINATION" in header else None
    descr_idx = header.index("DESCRIPTION") if "DESCRIPTION" in header else None
    product_cols = [(i, h) for i, h in enumerate(header)
                    if h and PRODUCT_CODE_RE.match(h)]
    return cc_idx, descr_idx, product_cols


def fix_one_sheet(ws, source_rows) -> list[dict]:
    """Apply formulas to ws. source_rows is the cached-value snapshot used to
    compute expected sums and to flatten ALL formulas to static values.

    Two passes:
      1. Flatten every cell whose value is a formula string ('=...') to the
         cached value from source_rows. openpyxl drops cached values on save,
         so any unflattened formula reads as None when re-loaded with
         data_only=True. This pass is global — covers SOHAND, costs,
         derived columns, everything.
      2. For every BEAD TOTALS row, write =SUM(top..bottom) in each product
         column (re-applying the corrected totals formula on top).
    """
    header = [str(c).strip() if c is not None else "" for c in source_rows[find_header_row_idx(source_rows)]]
    hdr_idx = find_header_row_idx(source_rows)
    cc_idx, descr_idx, product_cols = detect_columns(header)
    if descr_idx is None or not product_cols:
        return []

    actions = []

    # Pass 1 (global): flatten every formula cell to its cached value
    flattened = 0
    for r_idx in range(len(source_rows)):
        for c_idx in range(len(source_rows[r_idx])):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cached = source_rows[r_idx][c_idx]
                cell.value = cached
                flattened += 1

    # Pass 2: walk sections, write the corrected =SUM(top..bottom) formulas
    section_start_idx = None
    for r_idx in range(hdr_idx + 1, len(source_rows)):
        row = source_rows[r_idx]
        cc_val = row[cc_idx] if (cc_idx is not None and cc_idx < len(row)) else None
        descr_val = row[descr_idx] if descr_idx < len(row) else None
        cc_str = str(cc_val).strip() if cc_val else ""
        descr_str = str(descr_val).strip() if descr_val else ""

        if cc_str:
            section_start_idx = r_idx

        if "BEAD TOTALS" in descr_str.upper():
            if section_start_idx is None:
                continue
            totals_xlsx_row = r_idx + 1
            data_first = section_start_idx + 1
            data_last = r_idx
            for c_idx, h in product_cols:
                col_letter = get_column_letter(c_idx + 1)
                formula = f"=SUM({col_letter}{data_first}:{col_letter}{data_last})"
                ws.cell(row=totals_xlsx_row, column=c_idx + 1).value = formula
                expected = 0
                for rr in range(section_start_idx, r_idx):
                    v = source_rows[rr][c_idx] if c_idx < len(source_rows[rr]) else None
                    if isinstance(v, (int, float)):
                        expected += v
                actions.append({
                    "row": totals_xlsx_row,
                    "col": col_letter,
                    "header": h,
                    "formula": formula,
                    "expected": expected,
                    "section": f"{data_first}..{data_last}",
                })
            section_start_idx = None

    if flattened:
        actions.append({"_meta": True, "flattened_data_cells": flattened})
    return actions


def main():
    files = sorted(DATA_V2.glob("*.xlsx"))
    grand_total_actions = 0
    for f in files:
        print(f"\n=== {f.name} ===")
        # Open WITHOUT data_only so we don't strip formulas elsewhere
        wb = openpyxl.load_workbook(f, data_only=False)
        # Also open a cached-value snapshot for expected-sum calculation
        wb_cached = openpyxl.load_workbook(f, data_only=True)
        any_changes = False
        for sn in wb.sheetnames:
            ws = wb[sn]
            ws_cached = wb_cached[sn]
            cached_rows = list(ws_cached.iter_rows(values_only=True))
            actions = fix_one_sheet(ws, cached_rows)
            if not actions:
                continue
            print(f"  Sheet: {sn}")
            # Group by row for cleaner output
            meta = next((a for a in actions if a.get("_meta")), None)
            actions_real = [a for a in actions if not a.get("_meta")]
            by_row: dict[int, list] = {}
            for a in actions_real:
                by_row.setdefault(a["row"], []).append(a)
            for row, acts in sorted(by_row.items()):
                print(f"    Row {row}  (data {acts[0]['section']}):")
                for a in acts:
                    print(f"      {a['col']:>2} {a['header']:<11}  "
                          f"-> {a['formula']:<22}  expected={a['expected']}")
            if meta:
                print(f"    Flattened {meta['flattened_data_cells']} data-cell formulas to static values")
            any_changes = True
            grand_total_actions += len(actions_real)
        if any_changes:
            wb.save(f)
            print(f"  saved: {f.name}")
        else:
            print("  (no changes)")
    print(f"\nTotal cells updated: {grand_total_actions}")


if __name__ == "__main__":
    main()
